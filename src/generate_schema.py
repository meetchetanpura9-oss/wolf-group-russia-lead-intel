from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from pathlib import Path

# Paths relative to project root
data_dir = Path("c:/Users/meetc/wolf-group-russia-lead-intel/data")
docs_dir = Path("c:/Users/meetc/wolf-group-russia-lead-intel/docs")

xlsx_path = data_dir / "wolf_group_buyer_intelligence_schema.xlsx"
md_path = docs_dir / "business_requirements.md"

# Ensure directories exist
data_dir.mkdir(parents=True, exist_ok=True)
docs_dir.mkdir(parents=True, exist_ok=True)

# Excel schema
wb = Workbook()
ws = wb.active
ws.title = "Company Info"
schemas = {
    "Company Info": [
        "Company Name", "Legal Name", "Country", "City", "Region", "Website",
        "Company Type", "Product Category", "Product Focus", "Import Activity",
        "Company Size", "Market Relevance", "Recent Activity"
    ],
    "Contact Info": [
        "Company Name", "Business Email", "Business Phone", "WhatsApp",
        "Contact Person", "Designation", "Contact Page URL", "Contact Source URL"
    ],
    "Social Media": [
        "Company Name", "LinkedIn", "VK", "Telegram", "Instagram", "Facebook",
        "YouTube", "Other Social URL"
    ],
    "Verification": [
        "Company Name", "Source URL", "Source Type", "Date Verified",
        "Website Verified", "Phone Verified", "Email Verified",
        "WhatsApp Verified", "Decision Maker Verified", "Data Confidence",
        "Verification Status", "Notes"
    ],
}
for sheet, cols in schemas.items():
    ws = wb.create_sheet(sheet) if sheet != "Company Info" else wb["Company Info"]
    for col_idx, value in enumerate(cols, 1):
        cell = ws.cell(row=1, column=col_idx, value=value)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center")
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    for col_idx in range(1, len(cols) + 1):
        ws.column_dimensions[ws.cell(1, col_idx).column_letter].width = min(max(len(cols[col_idx-1]) + 3, 14), 28)

wb.save(xlsx_path)

# Markdown requirements document
md = """# Wolf Group Russia B2B Buyer Intelligence — Business Requirements

## 1. Product Understanding

- **Product categories:** Wolf Group India positions itself as a manufacturer and exporter of both **porcelain and ceramic tiles**, with a focus on durable, design-oriented flooring and surfaces. The official site explicitly lists Porcelain Tiles and Ceramic Tiles. Source: https://wolfgroupindia.com/
- **Formats:** The current 3D collection pages list **600×600 mm, 600×1200 mm, 800×1600 mm, 1200×1200 mm, and 1200×1800 mm** formats. These cover standard commercial sizes through large-format applications. Source: https://wolfporcelaintiles.com/3d-collections/
- **Finishes/design:** Published Wolf material describes **glossy, high-glossy, matt, carving, matt-rustic, and matt-satin** finishes, with porcelain collections also described as matte, polished, and textured. Source: https://wolfporcelaintiles.com/blog/tile-suppliers-in-morbi/
- **Price positioning:** Wolf's public material supports a **value-to-mid/premium B2B positioning**, rather than an ultra-luxury-only position: the company emphasizes quality, design, competitive Morbi manufacturing economics, and export supply. A current Wolf article describes Indian porcelain collections as competitive with European mid-premium catalogues. This should be treated as a market-positioning assessment, not a published official price list. Source: https://wolfporcelaintiles.com/india-vs-italy-vs-spain-tiles-for-polish-distributors/
- **Export footprint:** Wolf's current website says its products reach **40+ countries** and describes international customers including importers, distributors, retailers and architects. Public customer testimonials on the official site reference markets including **Yemen, UAE and Senegal**. Source: https://wolfporcelaintiles.com/ and https://wolfporcelaintiles.com/contact/

## 2. Buyer Personas

### 2.1 Importers — Priority 1
Importers buy in container-scale or other large B2B quantities and care strongly about landed cost, documentation, product consistency, logistics and supplier reliability. They can have longer qualification cycles because samples, specifications, compliance documents and commercial terms may need approval before a first shipment.

### 2.2 Distributors / Wholesalers — Priority 2
Distributors and wholesalers typically maintain inventory and supply multiple retailers or project customers. They value repeatability of supply, catalogue breadth, margin, reorder speed and geographic coverage; purchasing can become recurring once a supplier is approved.

### 2.3 Retailers / Showrooms — Priority 3
Retailers and showrooms buy selected ranges that fit local consumer demand and showroom space. Order sizes are generally smaller than direct import containers unless the retailer imports itself; decisions can be relatively fast when samples, pricing and availability are clear.

### 2.4 Construction / Developers — Priority 4
Developers and construction/project buyers purchase against project specifications, schedules and quantities. Order values can be large, but sales cycles are often slower because architects, procurement teams, contractors and project managers may all influence approval.

## 3. Lead Qualification and Scoring

Overall Lead Score = weighted score from six factors, normalized to 0–100.

| Factor | Weight | Measurement |
|---|---:|---|
| Product Fit | 25% | Score based on verified evidence that the company buys, distributes, sells or specifies ceramic/porcelain tiles and that Wolf's formats/finishes are relevant. |
| Import Activity | 25% | Score based on verified evidence of direct importing, international sourcing, customs/trade activity, import-oriented services or explicit overseas supplier relationships. |
| Company Size | 15% | Estimate from verifiable public evidence such as employee range, branch/network footprint, showroom/distribution scale and stated operations; classify as Small/Medium/Large/Unknown. |
| Market Relevance | 15% | Score based on Russia location, target geography, customer segment and fit with Wolf's intended B2B buyer personas. |
| Recent Activity | 10% | Score based on recent website updates, product launches, trade-fair participation, hiring, expansion, social activity or other dated public business signals. |
| Decision Maker Found | 10% | Score based on whether a relevant public business contact is identified, such as Procurement, Import, Purchasing, Commercial or Sales Director. |

### Lead Buckets

- **HOT:** 80–100
- **WARM:** 60–79
- **POTENTIAL:** 40–59
- **LOW:** 0–39

### Scoring discipline

The scoring model must use only verifiable evidence. If evidence is missing, the relevant field is **Unknown / Not Found** rather than an inferred fact. Unknown values must not be converted into positive evidence.

## 4. Final Excel / CSV Schema

The master workbook will contain four tabs:

### Company Info

- Company Name
- Legal Name
- Country
- City
- Region
- Website
- Company Type
- Product Category
- Product Focus
- Import Activity
- Company Size
- Market Relevance
- Recent Activity

### Contact Info

- Company Name
- Business Email
- Business Phone
- WhatsApp
- Contact Person
- Designation
- Contact Page URL
- Contact Source URL

### Social Media

- Company Name
- LinkedIn
- VK
- Telegram
- Instagram
- Facebook
- YouTube
- Other Social URL

### Verification

- Company Name
- Source URL
- Source Type
- Date Verified
- Website Verified
- Phone Verified
- Email Verified
- WhatsApp Verified
- Decision Maker Verified
- Data Confidence
- Verification Status
- Notes

## 5. Data-Quality Rules

1. **Website required:** A company cannot enter the final qualified-lead dataset without a publicly verifiable website or equivalent authoritative company page.
2. **Russia relevance required:** The company must have verifiable Russia-based operations, customers, branches or market activity relevant to the project.
3. **Source required:** Every important contact/intelligence field must have a source URL.
4. **Phone verification:** Do not record a phone number unless it is publicly displayed on a verifiable source.
5. **Email verification:** Do not infer or generate email addresses. Record only publicly listed business emails.
6. **WhatsApp rule:** Record WhatsApp only when the company explicitly presents that number as a business WhatsApp/contact channel. A phone number alone is not proof of WhatsApp.
7. **Decision-maker rule:** Record a person and designation only when the relationship to the company and role are supported by a public source.
8. **No hallucination:** If information cannot be verified, record `Unknown` or `Not Found`; never invent a value.
9. **Deduplication:** Duplicate companies must be consolidated using normalized company name, website domain and phone number, with fuzzy matching used as a secondary check.
10. **Website relevance:** The surviving company must be demonstrably relevant to tiles, building materials, import/distribution, retail/showroom activity or construction/project procurement.
11. **Date tracking:** Verification dates must be recorded so stale information can be identified later.
12. **Confidence:** Each record receives a confidence level based on source quality and completeness.
13. **Source hierarchy:** Prefer official company websites and first-party public profiles; use third-party directories/trade sources as supporting evidence and label the source type.
14. **Compliance:** Do not bypass authentication, paywalls, robots restrictions or access controls. Do not scrape search engines directly without an authorized API.
15. **Marketing separation:** Publicly discovered contact data is for intelligence/research; it is not automatically treated as marketing consent. Any outreach workflow must apply the project's consent and human-approval controls.

## 6. Research Source Notes

Primary product research sources:
- Wolf Group India official site: https://wolfgroupindia.com/
- Wolf Porcelain Tiles official site: https://wolfporcelaintiles.com/
- Wolf Porcelain Tiles 3D collections: https://wolfporcelaintiles.com/3d-collections/
- Wolf Porcelain Tiles contact page: https://wolfporcelaintiles.com/contact/
- Wolf Porcelain Tiles team page: https://wolfporcelaintiles.com/team/

## 7. Phase 1 Completion Criteria

Phase 1 is complete when:
- Product understanding is documented.
- Four buyer personas are defined.
- The weighted scoring model and measurement rules are fixed.
- The four-tab Excel schema is locked.
- Data-quality rules are documented.
- This requirements file is committed to GitHub.

This document is the reference point for all later collection, cleaning, qualification and reporting phases.
"""

md_path.write_text(md, encoding="utf-8")

print(f"Created: {xlsx_path}")
print(f"Created: {md_path}")
